import os
import re
from pathlib import Path
from openai import OpenAI
import openpyxl
from openpyxl import Workbook

try:
    import pymupdf as fitz
except ImportError:
    import fitz

# 初始化OpenAI客户端
client = OpenAI(
    api_key="sk-f8a5efd1bb8e466da74b59ac1b269099",
    base_url="https://dashscope.aliyuncs.com/compatible-mode/v1"
)

# 目标目录
target_dir = r"E:/Projects/Intelligent_Teaching_Assistant/课程论文"
output_filename = r"E:/Projects/Intelligent_Teaching_Assistant/课程论文-批改.xlsx"

def pdf_to_text(pdf_path, max_pages=None):
    """将PDF转换为文本"""
    try:
        doc = fitz.open(pdf_path)
        text_content = ""
        
        # 如果max_pages为None，处理所有页面
        page_count = len(doc) if max_pages is None else min(len(doc), max_pages)
        
        for page_num in range(page_count):
            page = doc[page_num]
            text_content += page.get_text()
            text_content += "\n\n"  # 页面间分隔
        
        doc.close()
        return text_content.strip()
    except Exception as e:
        print(f"  PDF转文本出错: {str(e)}")
        return None

def extract_student_info(filename):
    """从文件名提取学生姓名和学号"""
    # 文件名格式：姓名 学号.pdf
    name_match = re.match(r'^(.+?)\s+(\d+)\.pdf$', filename)
    if name_match:
        name = name_match.group(1).strip()
        student_id = name_match.group(2).strip()
        return name, student_id
    return None, None

def grade_paper(pdf_path):
    """使用通义千问批改论文"""
    try:
        # 将PDF转换为文本
        print(f"  正在提取PDF文本...")
        text_content = pdf_to_text(pdf_path)
        
        if not text_content:
            return None, "无法提取PDF文本内容"
        
        # 限制文本长度，避免超出API限制
        if len(text_content) > 8000:
            text_content = text_content[:8000] + "...[文本过长，已截断]"
        
        print(f"  正在调用AI批改...")
        
        # 构建消息
        response = client.chat.completions.create(
            model="qwen-max",
            messages=[
                {
                    "role": "system",
                    "content": "你是一位严谨的机械原理课程教师，正在批改学生的课程论文。请从以下几个方面进行评价：1.内容完整性和逻辑性 2.理论知识掌握程度 3.分析问题的深度 4.文字表达和格式规范性。给出70-100分之间的分数，必须是5的倍数。"
                },
                {
                    "role": "user",
                    "content": f"请批改以下机械原理课程论文，给出分数和评语：\n\n{text_content}\n\n请按以下格式回复：\n分数：XX\n评语：[详细评语，100-200字]"
                }
            ],
            stream=False
        )
        
        # 提取分数和评语
        content_text = response.choices[0].message.content
        
        # 提取分数
        score_match = re.search(r'分数[：:]\s*(\d+)', content_text)
        if not score_match:
            score_match = re.search(r'(\d+)', content_text)
        
        # 提取评语
        comment_match = re.search(r'评语[：:]\s*(.+)', content_text, re.DOTALL)
        comment = comment_match.group(1).strip() if comment_match else content_text
        
        if score_match:
            score = int(score_match.group(1))
            # 确保分数在70-100之间且是5的倍数
            score = max(70, min(100, score))
            score = round(score / 5) * 5
            return score, comment
        else:
            return 80, comment  # 默认分数
            
    except Exception as e:
        print(f"  批改出错: {str(e)}")
        return None, str(e)

def main():
    # 获取所有PDF文件
    pdf_files = list(Path(target_dir).glob("*.pdf"))
    
    # 按学号排序
    def extract_student_id(file_path):
        name, student_id = extract_student_info(file_path.name)
        return student_id if student_id else "999999999"
    
    pdf_files.sort(key=extract_student_id)
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "论文批改结果"
    
    # 设置表头
    ws['A1'] = "序号"
    ws['B1'] = "学生姓名"
    ws['C1'] = "学号"
    ws['D1'] = "文件名"
    ws['E1'] = "分数"
    ws['F1'] = "评语"
    
    print(f"开始批改论文，共 {len(pdf_files)} 份...")
    
    # 批改每份论文
    for idx, pdf_file in enumerate(pdf_files, start=1):
        print(f"\n正在批改 [{idx}/{len(pdf_files)}]: {pdf_file.name}")
        
        # 提取学生信息
        name, student_id = extract_student_info(pdf_file.name)
        
        score, comment = grade_paper(pdf_file)
        
        ws[f'A{idx+1}'] = idx
        ws[f'B{idx+1}'] = name if name else "未知"
        ws[f'C{idx+1}'] = student_id if student_id else "未知"
        ws[f'D{idx+1}'] = pdf_file.name
        ws[f'E{idx+1}'] = score if score else "批改失败"
        ws[f'F{idx+1}'] = comment if comment else "无评语"
        
        if score:
            print(f"学生: {name} ({student_id})")
            print(f"分数: {score}")
            print(f"评语: {comment[:50]}..." if len(comment) > 50 else f"评语: {comment}")
        
        # 每批改3份保存一次，防止数据丢失（论文批改较慢）
        if idx % 3 == 0:
            wb.save("论文批改结果_临时.xlsx")
            print("已保存临时结果")
    
    # 保存最终结果
    output_file = output_filename
    wb.save(output_file)
    
    # 删除临时文件
    temp_file = Path("论文批改结果_临时.xlsx")
    if temp_file.exists():
        temp_file.unlink()
    
    print(f"\n批改完成！结果已保存到: {output_file}")

if __name__ == "__main__":
    main()