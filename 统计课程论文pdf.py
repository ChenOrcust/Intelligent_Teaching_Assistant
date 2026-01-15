import os
import re
from pathlib import Path
import openpyxl
from openpyxl import Workbook

# 目标目录
target_dir = r"E:/Document/机械原理2025-2026/课程论文"

# 创建工作簿
wb = Workbook()
ws = wb.active
ws.title = "PDF文件统计"

# 设置表头
ws['A1'] = "序号"
ws['B1'] = "文件名"

# 获取所有PDF文件
pdf_files = []
for file in Path(target_dir).glob("*.pdf"):
    pdf_files.append(file)

# 按学号排序
def extract_student_id(file_path):
    # 从文件名中提取学号（连续的数字）
    match = re.search(r'(\d{10,})', file_path.name)
    return int(match.group(1)) if match else 0

pdf_files.sort(key=extract_student_id)

# 写入数据
for idx, pdf_file in enumerate(pdf_files, start=1):
    ws[f'A{idx+1}'] = idx
    ws[f'B{idx+1}'] = pdf_file.name

# 保存文件
output_file = "课程论文PDF文件统计.xlsx"
wb.save(output_file)

print(f"统计完成！共找到 {len(pdf_files)} 个PDF文件")
print(f"结果已保存到: {output_file}")
