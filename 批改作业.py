import os
import re
from pathlib import Path
from openai import OpenAI
import openpyxl
from openpyxl import Workbook
import base64
from io import BytesIO
from PIL import Image

try:
    import pymupdf as fitz
except ImportError:
    import fitz

# 初始化OpenAI客户端
client = OpenAI(
    api_key="sk-f8a5efd1bb8e466da74b59ac1b269099",
    base_url="https://dashscope.aliyuncs.com/compatible-mode/v1"
)

# 目标根目录
root_dir = r"E:/Projects/Intelligent_Teaching_Assistant/小组作业/二人小组作业"
output_filename = r"E:/Projects/Intelligent_Teaching_Assistant/二人小组作业-批改.xlsx"

# 排除的路径列表（已批改的章节）
excluded_paths = [
    "第2章",  # 示例：排除第2章
    # 可以在这里添加更多要排除的路径
]

def pdf_to_images(pdf_path, max_pages=10):
    """将PDF转换为图片列表"""
    images = []
    doc = fitz.open(pdf_path)
    
    # 如果max_pages为None，处理所有页面
    page_count = len(doc) if max_pages is None else min(len(doc), max_pages)
    
    for page_num in range(page_count):
        page = doc[page_num]
        # 设置缩放比例，提高图片质量
        mat = fitz.Matrix(2, 2)
        pix = page.get_pixmap(matrix=mat)
        
        # 转换为PIL Image
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        
        # 转换为base64
        buffered = BytesIO()
        img.save(buffered, format="JPEG", quality=85)
        img_base64 = base64.b64encode(buffered.getvalue()).decode('utf-8')
        images.append(img_base64)
    
    doc.close()
    return images

def grade_assignment(pdf_path):
    """使用通义千问批改作业"""
    try:
        # 将PDF转换为图片
        print(f"  正在转换PDF为图片...")
        images = pdf_to_images(pdf_path, max_pages=None)  # None表示处理所有页面
        
        # 构建消息内容
        content = [
            {
                "type": "text",
                "text": "这是一份机械原理课程的作业，请仔细评阅。从内容完整性、准确性、规范性等方面进行评分。给出80-100分之间的分数，必须是5的倍数（80、85、90、95、100）。请按以下格式回复：\n分数：XX\n评语：[简短评语，不超过50字]"
            }
        ]
        
        # 添加所有页面的图片
        for img_base64 in images:
            content.append({
                "type": "image_url",
                "image_url": {
                    "url": f"data:image/jpeg;base64,{img_base64}"
                }
            })
        
        print(f"  正在调用AI批改...")
        # 构建消息
        response = client.chat.completions.create(
            model="qwen-vl-max",
            messages=[
                {
                    "role": "system",
                    "content": "你是一位严谨的机械原理课程教师。"
                },
                {
                    "role": "user",
                    "content": content
                }
            ],
            stream=False
        )
        
        # 提取分数和评语
        content_text = response.choices[0].message.content
        
        # 提取分数
        score_match = re.search(r'分数[：:]\s*(\d+)', content_text)
        if not score_match:
            score_match = re.search(r'(\d+)', content_text)
        
        # 提取评语
        comment_match = re.search(r'评语[：:]\s*(.+)', content_text, re.DOTALL)
        comment = comment_match.group(1).strip() if comment_match else content_text
        
        if score_match:
            score = int(score_match.group(1))
            # 确保分数在80-100之间且是5的倍数
            score = max(80, min(100, score))
            score = round(score / 5) * 5
            return score, comment
        else:
            return 85, comment  # 默认分数
            
    except Exception as e:
        print(f"  批改出错: {str(e)}")
        return None, str(e)

def process_folder(folder_path, ws, start_row=2, total_groups=None, wb=None):
    """处理单个文件夹中的PDF文件"""
    # 获取所有PDF文件
    pdf_files = list(Path(folder_path).glob("*.pdf"))
    
    # 按小组序号排序
    def extract_group_number(file_path):
        match = re.search(r'第(\d+)小组', file_path.name)
        return int(match.group(1)) if match else 999
    
    pdf_files.sort(key=extract_group_number)
    
    # 创建已提交小组的集合
    submitted_groups = set()
    for pdf_file in pdf_files:
        group_num = extract_group_number(pdf_file)
        if group_num != 999:
            submitted_groups.add(group_num)
    
    print(f"处理文件夹: {folder_path.name}，共 {len(pdf_files)} 份作业...")
    
    current_row = start_row
    
    # 如果指定了总组数，按组号顺序处理
    if total_groups:
        for group_num in range(1, total_groups + 1):
            # 查找该组的PDF文件
            group_file = None
            for pdf_file in pdf_files:
                if extract_group_number(pdf_file) == group_num:
                    group_file = pdf_file
                    break
            
            if group_file:
                # 有提交，正常批改
                print(f"\n正在批改第{group_num}小组: {group_file.name}")
                score, comment = grade_assignment(group_file)
                
                ws[f'A{current_row}'] = group_num
                ws[f'B{current_row}'] = group_file.name
                ws[f'C{current_row}'] = score if score else "批改失败"
                ws[f'D{current_row}'] = comment if comment else "无评语"
                
                if score:
                    print(f"分数: {score}")
                    print(f"评语: {comment[:30]}..." if len(comment) > 30 else f"评语: {comment}")
            else:
                # 没有提交，标记为0分
                print(f"\n第{group_num}小组未提交作业")
                ws[f'A{current_row}'] = group_num
                ws[f'B{current_row}'] = f"第{group_num}小组（未提交）"
                ws[f'C{current_row}'] = 0
                ws[f'D{current_row}'] = "未提交作业"
            
            current_row += 1
            
            # 每处理5个小组保存一次
            if group_num % 5 == 0 and wb:
                wb.save(output_filename)
                print("已保存临时结果")
    else:
        # 不指定总组数，只批改已提交的
        for idx, pdf_file in enumerate(pdf_files, start=1):
            print(f"\n正在批改 [{idx}/{len(pdf_files)}]: {pdf_file.name}")
            
            score, comment = grade_assignment(pdf_file)
            
            ws[f'A{current_row}'] = idx
            ws[f'B{current_row}'] = pdf_file.name
            ws[f'C{current_row}'] = score if score else "批改失败"
            ws[f'D{current_row}'] = comment if comment else "无评语"
            
            if score:
                print(f"分数: {score}")
                print(f"评语: {comment[:30]}..." if len(comment) > 30 else f"评语: {comment}")
            
            current_row += 1
            
            # 每批改5份保存一次
            if idx % 5 == 0 and wb:
                wb.save(output_filename)
                print("已保存临时结果")
    
    return current_row

def main():
    """主函数"""
    # 获取根目录下的所有子文件夹
    root_path = Path(root_dir)
    if not root_path.exists():
        print(f"错误：目录不存在 {root_dir}")
        return
    
    # 获取所有章节文件夹
    chapter_folders = [f for f in root_path.iterdir() if f.is_dir()]
    
    # 过滤掉排除的路径
    chapter_folders = [f for f in chapter_folders if not any(excluded in f.name for excluded in excluded_paths)]
    
    # 按章节序号排序
    def extract_chapter_number(folder_path):
        match = re.search(r'第(\d+)章', folder_path.name)
        return int(match.group(1)) if match else 999
    
    chapter_folders.sort(key=extract_chapter_number)
    
    if not chapter_folders:
        print("没有找到需要批改的章节文件夹")
        return
    
    print(f"找到 {len(chapter_folders)} 个章节需要批改")
    for folder in chapter_folders:
        print(f"  - {folder.name}")
    
    # 询问总组数
    total_groups_input = input("\n请输入总共有多少个小组（直接回车跳过，只批改已提交的）: ").strip()
    total_groups = None
    if total_groups_input:
        try:
            total_groups = int(total_groups_input)
            print(f"将按 {total_groups} 个小组处理，未提交的标记为0分")
        except ValueError:
            print("输入无效，将只批改已提交的作业")
    
    # 创建或加载工作簿
    output_path = Path(output_filename)
    if output_path.exists():
        print(f"\n检测到已存在的文件: {output_filename}")
        wb = openpyxl.load_workbook(output_filename)
    else:
        wb = Workbook()
        # 删除默认的Sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
    
    # 处理每个章节
    for folder in chapter_folders:
        chapter_name = folder.name
        
        # 创建或获取工作表
        if chapter_name in wb.sheetnames:
            print(f"\n工作表 '{chapter_name}' 已存在，跳过...")
            continue
        
        ws = wb.create_sheet(title=chapter_name)
        
        # 设置表头
        ws['A1'] = "小组序号"
        ws['B1'] = "文件名"
        ws['C1'] = "分数"
        ws['D1'] = "评语"
        
        # 处理该章节的作业
        process_folder(folder, ws, start_row=2, total_groups=total_groups, wb=wb)
        
        # 保存当前进度
        wb.save(output_filename)
        print(f"\n{chapter_name} 批改完成")
    
    # 保存最终结果
    wb.save(output_filename)
    print(f"\n所有批改完成！结果已保存到: {output_filename}")

if __name__ == "__main__":
    main()
