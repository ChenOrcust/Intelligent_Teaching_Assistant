import os
import re
from pathlib import Path
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# 目标目录
target_dir = r"E:\Document\机械原理2025-2026\小组作业\二人小组作业"

# 创建工作簿
wb = Workbook()
# 删除默认的sheet
wb.remove(wb.active)

# 获取所有子文件夹
subdirs = [d for d in Path(target_dir).iterdir() if d.is_dir()]
subdirs.sort()

total_folders = 0
total_files = 0

# 遍历每个子文件夹
for subdir in subdirs:
    folder_name = subdir.name
    
    # 获取该文件夹下的所有PDF文件
    pdf_files = list(subdir.glob("*.pdf"))
    
    if pdf_files:  # 只为有PDF文件的文件夹创建工作表
        # 创建工作表，使用文件夹名作为sheet名（Excel sheet名有长度限制）
        sheet_name = folder_name[:31]  # Excel工作表名最多31个字符
        ws = wb.create_sheet(title=sheet_name)
        
        # 设置表头
        ws['A1'] = "序号"
        ws['B1'] = "文件名"
        
        # 按小组序号排序
        def extract_group_number(file_path):
            # 从文件名中提取"第X小组"的数字
            match = re.search(r'第(\d+)小组', file_path.name)
            return int(match.group(1)) if match else 999
        
        pdf_files.sort(key=extract_group_number)
        
        # 写入数据
        for idx, pdf_file in enumerate(pdf_files, start=1):
            ws[f'A{idx+1}'] = idx
            ws[f'B{idx+1}'] = pdf_file.name
        
        total_folders += 1
        total_files += len(pdf_files)
        print(f"文件夹 [{folder_name}] 找到 {len(pdf_files)} 个PDF文件")

# 保存文件
output_file = "二人小组作业PDF统计.xlsx"
wb.save(output_file)

print(f"\n统计完成！")
print(f"共处理 {total_folders} 个文件夹")
print(f"共找到 {total_files} 个PDF文件")
print(f"结果已保存到: {output_file}")
