import openpyxl
from pathlib import Path
from datetime import datetime
import sys

# 文件路径
student_list_file = r"E:\Projects\Intelligent_Teaching_Assistant\2024-2025-秋-302074030-02-机械原理-选课学生名单-2025-09-28.xls"
group_file = r"E:\Projects\Intelligent_Teaching_Assistant\分组-2025-2026-秋-302074030-02-机械原理.xlsx"
grade_file = r"E:\Projects\Intelligent_Teaching_Assistant\二人小组作业-批改.xlsx"

# 默认输出文件名（不带时间戳）
default_output_file = r"E:\Projects\Intelligent_Teaching_Assistant\二人小组作业-成绩整理.xlsx"

def normalize_student_id(student_id):
    """标准化学号格式，去除空格和特殊字符"""
    if student_id is None:
        return ""
    student_id_str = str(student_id).strip()
    # 去除可能的.0后缀（Excel数字格式）
    if student_id_str.endswith('.0'):
        student_id_str = student_id_str[:-2]
    return student_id_str

def load_student_list(file_path):
    """加载学生选课名单，从B8开始读取学号，C8开始读取姓名"""
    print("正在加载学生选课名单...")
    
    # 使用openpyxl读取xls文件（需要先转换为xlsx）
    # 或者使用xlrd库
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"无法直接读取.xls文件，尝试使用xlrd...")
        # 如果是.xls文件，需要先转换
        import xlrd
        import openpyxl as oxl
        
        # 读取xls
        xls_book = xlrd.open_workbook(file_path)
        xls_sheet = xls_book.sheet_by_index(0)
        
        # 创建新的xlsx工作簿
        wb = oxl.Workbook()
        ws = wb.active
        
        # 复制数据
        for row_idx in range(xls_sheet.nrows):
            for col_idx in range(xls_sheet.ncols):
                ws.cell(row=row_idx+1, column=col_idx+1, value=xls_sheet.cell_value(row_idx, col_idx))
    
    ws = wb.active
    
    # 从第8行开始读取
    students = []
    for row in range(8, ws.max_row + 1):
        student_id = ws[f'B{row}'].value  # B列
        name = ws[f'C{row}'].value  # C列
        
        # 跳过空行
        if student_id is None or name is None:
            continue
        
        # 转换为字符串并去除空格
        student_id_str = normalize_student_id(student_id)
        name_str = str(name).strip()
        
        if student_id_str and name_str:
            students.append({
                'student_id': student_id_str,
                'name': name_str
            })
    
    wb.close()
    print(f"共加载 {len(students)} 名学生")
    return students

def load_group_info(file_path):
    """加载分组信息，读取二人小组工作表"""
    print("正在加载分组信息...")
    
    wb = openpyxl.load_workbook(file_path)
    
    # 读取二人小组工作表
    sheet_name = None
    for name in wb.sheetnames:
        if '二人小组' in name:
            sheet_name = name
            break
    
    if sheet_name is None:
        print("错误：找不到包含'二人小组'的工作表")
        print(f"可用的工作表: {wb.sheetnames}")
        wb.close()
        return {}
    
    ws = wb[sheet_name]
    print(f"使用工作表: {sheet_name}")
    
    # 创建学号到组号的映射
    student_to_group = {}
    
    # 当前组号（用于处理空组号的情况）
    current_group_num = None
    
    # 从第2行开始读取（第1行是表头）
    for row in range(2, ws.max_row + 1):
        group_num_cell = ws[f'A{row}'].value  # 第一列是组号
        
        # 如果当前行有组号，更新当前组号
        if group_num_cell is not None:
            try:
                current_group_num = int(group_num_cell)
            except (ValueError, TypeError):
                # 如果无法转换为整数，跳过这一行
                continue
        
        # 如果没有当前组号（表头行或无效行），跳过
        if current_group_num is None:
            continue
        
        # 遍历该行的所有学号列（从B列开始）
        has_student = False
        for col in range(2, ws.max_column + 1):
            col_letter = openpyxl.utils.get_column_letter(col)
            student_id = ws[f'{col_letter}{row}'].value
            
            if student_id is not None:
                student_id_str = normalize_student_id(student_id)
                if student_id_str:
                    student_to_group[student_id_str] = current_group_num
                    has_student = True
        
        # 如果这一行完全没有学号，检查是否是完全空行
        if not has_student:
            # 检查这一行是否完全为空（所有列都是None）
            is_empty_row = True
            for col in range(1, ws.max_column + 1):
                col_letter = openpyxl.utils.get_column_letter(col)
                if ws[f'{col_letter}{row}'].value is not None:
                    is_empty_row = False
                    break
            
            # 如果是完全空行，重置当前组号
            if is_empty_row:
                current_group_num = None
    
    wb.close()
    print(f"共加载 {len(student_to_group)} 名学生的分组信息")
    return student_to_group

def load_grades(file_path):
    """加载批改结果，读取所有章节的成绩"""
    print("正在加载批改结果...")
    
    wb = openpyxl.load_workbook(file_path)
    
    # 存储每个章节的成绩，格式：{章节名: {组号: 分数}}
    grades_by_chapter = {}
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        chapter_grades = {}
        
        # 从第2行开始读取（第1行是表头）
        for row in range(2, ws.max_row + 1):
            group_num = ws[f'A{row}'].value  # 小组序号
            score = ws[f'C{row}'].value  # 分数
            
            if group_num is not None:
                try:
                    group_num = int(group_num)
                    chapter_grades[group_num] = score if score is not None else 0
                except (ValueError, TypeError):
                    continue
        
        grades_by_chapter[sheet_name] = chapter_grades
        print(f"  - {sheet_name}: {len(chapter_grades)} 个小组")
    
    wb.close()
    return grades_by_chapter

def create_output(students, student_to_group, grades_by_chapter, output_path):
    """创建输出文件，按学生顺序排列成绩"""
    print("\n正在生成输出文件...")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "二人小组作业成绩"
    
    # 设置表头
    ws['A1'] = "序号"
    ws['B1'] = "学号"
    ws['C1'] = "姓名"
    ws['D1'] = "二人小组组号"
    
    # 获取所有章节名称并排序
    chapter_names = sorted(grades_by_chapter.keys())
    
    # 添加章节列
    for idx, chapter_name in enumerate(chapter_names, start=5):
        col_letter = openpyxl.utils.get_column_letter(idx)
        ws[f'{col_letter}1'] = chapter_name
    
    # 统计信息
    students_with_group = 0
    students_without_group = 0
    
    # 填充学生数据
    for row_idx, student in enumerate(students, start=2):
        student_id = student['student_id']
        name = student['name']
        
        # 获取学生的组号
        group_num = student_to_group.get(student_id, None)
        
        if group_num:
            students_with_group += 1
        else:
            students_without_group += 1
        
        # 填充基本信息
        ws[f'A{row_idx}'] = row_idx - 1  # 序号
        ws[f'B{row_idx}'] = student_id
        ws[f'C{row_idx}'] = name
        ws[f'D{row_idx}'] = group_num if group_num else "未分组"
        
        # 填充各章节成绩
        if group_num:
            for col_idx, chapter_name in enumerate(chapter_names, start=5):
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                score = grades_by_chapter[chapter_name].get(group_num, "")
                ws[f'{col_letter}{row_idx}'] = score
        else:
            # 未分组的学生，成绩留空
            for col_idx in range(5, 5 + len(chapter_names)):
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                ws[f'{col_letter}{row_idx}'] = ""
    
    # 保存文件
    wb.save(output_path)
    print(f"\n成绩整理完成！已保存到: {output_path}")
    print(f"共整理 {len(students)} 名学生，{len(chapter_names)} 个章节的成绩")
    print(f"  - 已分组学生: {students_with_group} 人")
    print(f"  - 未分组学生: {students_without_group} 人")

def main():
    """主函数"""
    print("=" * 60)
    print("二人小组作业成绩整理工具")
    print("=" * 60)
    print("\n功能说明：")
    print("1. 读取学生选课名单（学号和姓名）")
    print("2. 读取分组信息（二人小组）")
    print("3. 读取批改结果（各章节成绩）")
    print("4. 按学生选课名单顺序输出成绩表")
    print("\n" + "=" * 60 + "\n")
    
    # 询问是否使用固定文件名
    use_timestamp = input("是否在输出文件名中添加时间戳？(y/n，默认y): ").strip().lower()
    
    if use_timestamp == 'n':
        output_file = default_output_file
        # 检查文件是否存在
        if Path(output_file).exists():
            overwrite = input(f"文件 {output_file} 已存在，是否覆盖？(y/n): ").strip().lower()
            if overwrite != 'y':
                print("操作已取消")
                return
    else:
        # 生成带时间戳的输出文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = rf"E:\Projects\Intelligent_Teaching_Assistant\二人小组作业-成绩整理_{timestamp}.xlsx"
    
    # 1. 加载学生选课名单
    students = load_student_list(student_list_file)
    
    # 2. 加载分组信息
    student_to_group = load_group_info(group_file)
    
    # 3. 加载批改结果
    grades_by_chapter = load_grades(grade_file)
    
    # 4. 生成输出文件
    create_output(students, student_to_group, grades_by_chapter, output_file)
    
    print("\n处理完成！")
    print(f"输出文件: {output_file}")

if __name__ == "__main__":
    main()
